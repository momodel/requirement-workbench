from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

from ..config import AppSettings, DEFAULT_SETTINGS
from ..models import ProviderIssue
from .docling_normalizer import DoclingNormalizer


@dataclass(slots=True)
class NormalizedSource:
    source_kind: str
    normalize_status: str
    normalize_summary: str
    normalized_path: str | None = None
    index_input_mode: str | None = None


class SourceIngestionService:
    TEXT_FILE_SUFFIXES = {".md", ".markdown", ".txt"}
    SPREADSHEET_SUFFIXES = {".xlsx"}

    def __init__(
        self,
        settings: AppSettings = DEFAULT_SETTINGS,
        *,
        docling_normalizer: DoclingNormalizer | None = None,
    ):
        self.settings = settings
        self.docling_normalizer = docling_normalizer or DoclingNormalizer()

    def project_source_dir(self, project_id: str) -> Path:
        path = self.settings.projects_dir / project_id / "sources"
        path.mkdir(parents=True, exist_ok=True)
        return path

    @staticmethod
    def _summarize_text(text: str, *, fallback: str) -> str:
        summary = text.strip().replace("\n", " ")[:240]
        return summary or fallback

    @staticmethod
    def _infer_source_kind(suffix: str) -> str:
        if suffix in DoclingNormalizer.IMAGE_SUFFIXES:
            return "image"
        if suffix in DoclingNormalizer.AUDIO_SUFFIXES:
            return "audio"
        if suffix in {".md", ".markdown"}:
            return "markdown"
        if suffix == ".txt":
            return "text"
        return suffix.lstrip(".") or "file"

    def _write_normalized_markdown(self, source_dir: Path, raw_path: Path, text: str) -> Path:
        normalized_path = source_dir / f"{raw_path.stem}.normalized.md"
        normalized_path.write_text(text, encoding="utf-8")
        return normalized_path

    def ingest_text(self, project_id: str, name: str, text_content: str) -> tuple[str | None, NormalizedSource]:
        source_dir = self.project_source_dir(project_id)
        raw_path = source_dir / f"{name}.txt"
        raw_path.write_text(text_content, encoding="utf-8")
        summary = self._summarize_text(text_content, fallback=f"{name} 已入库。")
        return str(raw_path), NormalizedSource(
            source_kind="text",
            normalize_status="parsed",
            normalize_summary=summary,
            normalized_path=str(raw_path),
            index_input_mode="direct_text",
        )

    def ingest_url(self, project_id: str, name: str, source_url: str) -> tuple[str | None, NormalizedSource]:
        source_dir = self.project_source_dir(project_id)
        raw_path = source_dir / f"{name}.url.txt"
        raw_path.write_text(source_url, encoding="utf-8")
        parsed = urlparse(source_url)
        summary = (
            f"URL 已记录：{parsed.netloc}{parsed.path or '/'}。"
            "当前版本还没有抓取到页面正文；生成 normalized text 前不会进入项目知识库。"
        )
        return str(raw_path), NormalizedSource(
            source_kind="url",
            normalize_status="pending",
            normalize_summary=summary,
            normalized_path=None,
            index_input_mode=None,
        )

    def ingest_file(self, project_id: str, filename: str, file_bytes: bytes) -> tuple[str, NormalizedSource]:
        source_dir = self.project_source_dir(project_id)
        raw_path = source_dir / filename
        raw_path.write_bytes(file_bytes)

        suffix = raw_path.suffix.lower()
        source_kind = self._infer_source_kind(suffix)
        if suffix in DoclingNormalizer.AUDIO_SUFFIXES:
            return str(raw_path), NormalizedSource(
                source_kind="audio",
                normalize_status="processing",
                normalize_summary=f"{filename} 已入库，正在转写；完成后会自动进入项目知识库。",
                normalized_path=None,
                index_input_mode=None,
            )

        normalized_path = None
        summary = f"{filename} 已入库，等待文本标准化。"
        import_mode = None

        try:
            if suffix in self.TEXT_FILE_SUFFIXES:
                text = raw_path.read_text(encoding="utf-8", errors="ignore")
                normalized_path = raw_path
                summary = self._summarize_text(text, fallback=summary)
                import_mode = "direct_text"
            elif suffix in self.SPREADSHEET_SUFFIXES:
                workbook = load_workbook(str(raw_path), read_only=True, data_only=True)
                lines: list[str] = []
                for sheet in workbook.worksheets:
                    lines.append(f"# Sheet: {sheet.title}")
                    rows = list(sheet.iter_rows(values_only=True, max_row=6))
                    if rows:
                        header = [str(cell or "") for cell in rows[0]]
                        lines.append("表头: " + " | ".join(header))
                    for sample_row in rows[1:4]:
                        lines.append("样例: " + " | ".join(str(cell or "") for cell in sample_row))
                    lines.append(f"行数估计: {sheet.max_row}, 列数估计: {sheet.max_column}")
                text = "\n".join(lines)
                normalized_path = self._write_normalized_markdown(source_dir, raw_path, text)
                summary = self._summarize_text(text, fallback="XLSX 已转成摘要文本。")
                import_mode = "normalized_text"
            elif self.docling_normalizer.supports(raw_path):
                text = self.docling_normalizer.normalize_to_markdown(raw_path)
                normalized_path = self._write_normalized_markdown(source_dir, raw_path, text)
                summary = self._summarize_text(text, fallback=f"{filename} 已通过 Docling 转成文本。")
                import_mode = "normalized_text"
            else:
                raise ProviderIssue(
                    provider="SOURCE_INGESTION",
                    message=(
                        f"{filename} 当前还没有接入正式的 text-first 标准化链路，"
                        "因此不能标记成已解析或可索引。"
                    ),
                )
        except Exception as exc:
            message = exc.message if isinstance(exc, ProviderIssue) else str(exc)
            return str(raw_path), NormalizedSource(
                source_kind=source_kind,
                normalize_status="failed",
                normalize_summary=message,
                normalized_path=None,
                index_input_mode=None,
            )

        return str(raw_path), NormalizedSource(
            source_kind=source_kind,
            normalize_status="parsed",
            normalize_summary=summary,
            normalized_path=str(normalized_path) if normalized_path else None,
            index_input_mode=import_mode,
        )
